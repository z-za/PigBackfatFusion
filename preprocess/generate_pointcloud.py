# -*- coding: utf-8 -*-
"""
Generate segmented RGB images, segmented depth images, and pig-back point clouds
from RGB-D images and YOLO segmentation masks.
"""

import argparse
import json
from pathlib import Path

import cv2
import numpy as np
from openpyxl import load_workbook


DEFAULT_INTRINSICS = {
    "fx": 610.855224609375,
    "fy": 610.863891601562,
    "cx": 644.7053833007812,
    "cy": 362.82342529296875,
}

DEFAULT_DEPTH_SCALE_M = 0.001


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Generate segmented RGB images, segmented depth images, "
            "and point clouds from RGB-D images and YOLO masks."
        )
    )

    parser.add_argument(
        "--root_dir",
        type=str,
        required=True,
        help="Root directory of the dataset, e.g., F:/data"
    )
    parser.add_argument(
        "--excel_path",
        type=str,
        default=None,
        help="Path to data.xlsx. If not provided, root_dir/data.xlsx will be used."
    )

    parser.add_argument("--fx", type=float, default=DEFAULT_INTRINSICS["fx"])
    parser.add_argument("--fy", type=float, default=DEFAULT_INTRINSICS["fy"])
    parser.add_argument("--cx", type=float, default=DEFAULT_INTRINSICS["cx"])
    parser.add_argument("--cy", type=float, default=DEFAULT_INTRINSICS["cy"])

    parser.add_argument(
        "--depth_scale_m",
        type=float,
        default=DEFAULT_DEPTH_SCALE_M,
        help="Scale factor used to convert depth values to metres. Use 0.001 if depth is stored in millimetres."
    )
    parser.add_argument(
        "--min_depth_m",
        type=float,
        default=0.05,
        help="Minimum valid depth in metres."
    )
    parser.add_argument(
        "--max_depth_m",
        type=float,
        default=1.55,
        help="Maximum valid depth in metres."
    )

    parser.add_argument(
        "--auto_resize_rgb_to_depth",
        action="store_true",
        help="Resize RGB images to the depth image size if their resolutions are inconsistent."
    )
    parser.add_argument(
        "--skip_existing",
        action="store_true",
        help="Skip output files that already exist."
    )
    parser.add_argument(
        "--rgb_background",
        type=str,
        default="black",
        choices=["black", "white"],
        help="Background color for segmented RGB images."
    )

    return parser.parse_args()


def cell_to_str(v, pad_len=6):
    if v is None:
        return None

    s = str(v).strip()
    if s == "":
        return None

    if isinstance(v, float):
        if v.is_integer():
            s = str(int(v))
        else:
            s = s.rstrip("0").rstrip(".")

    if s.isdigit() and pad_len is not None:
        s = s.zfill(pad_len)

    return s


def parse_excel_pairs(excel_path):
    """
    Excel rule:
    - Odd columns, row 1: first-level folder name, usually date.
    - Odd columns, rows 2 onward: second-level folder names.
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    pairs = []
    seen = set()

    for col in range(1, ws.max_column + 1, 2):
        lvl1 = cell_to_str(ws.cell(row=1, column=col).value, pad_len=6)
        if not lvl1:
            continue

        for row in range(2, ws.max_row + 1):
            lvl2 = cell_to_str(ws.cell(row=row, column=col).value, pad_len=6)
            if not lvl2:
                continue

            key = (lvl1, lvl2)
            if key not in seen:
                seen.add(key)
                pairs.append(key)

    return pairs


def cv_imread(path, flags=cv2.IMREAD_UNCHANGED):
    path = Path(path)
    data = np.fromfile(str(path), dtype=np.uint8)
    if data.size == 0:
        return None
    return cv2.imdecode(data, flags)


def cv_imwrite(path, img):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    ext = path.suffix.lower()
    if ext == "":
        ext = ".png"
        path = path.with_suffix(ext)

    ok, buf = cv2.imencode(ext, img)
    if not ok:
        raise RuntimeError(f"Failed to save image: {path}")

    buf.tofile(str(path))


def get_image_map(folder):
    exts = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"}
    folder = Path(folder)

    if not folder.exists():
        return {}

    image_map = {}
    for p in folder.iterdir():
        if p.is_file() and p.suffix.lower() in exts:
            image_map[p.stem] = p

    return image_map


def find_calibration_json(base_dir, root_dir):
    base_dir = Path(base_dir)
    root_dir = Path(root_dir)

    candidates = [
        base_dir / "camera_calibration.json",
        base_dir.parent / "camera_calibration.json",
        root_dir / "camera_calibration.json",
    ]

    for p in candidates:
        if p.exists():
            return p

    return None


def parse_intrinsics_from_json(json_path):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    intr = None
    depth_scale_m = None

    candidate_dicts = []

    if isinstance(data, dict):
        candidate_dicts.append(data)
        for key in [
            "intrinsics",
            "depth_intrinsics",
            "color_intrinsics",
            "camera_intrinsics",
            "depth_camera_intrinsics",
        ]:
            if key in data and isinstance(data[key], dict):
                candidate_dicts.append(data[key])

    def try_pick(d):
        keys = {k.lower(): k for k in d.keys()}
        if all(k in keys for k in ["fx", "fy", "cx", "cy"]):
            return {
                "fx": float(d[keys["fx"]]),
                "fy": float(d[keys["fy"]]),
                "cx": float(d[keys["cx"]]),
                "cy": float(d[keys["cy"]]),
            }
        return None

    for d in candidate_dicts:
        intr = try_pick(d)
        if intr is not None:
            break

    def search_key(obj, key_name):
        if isinstance(obj, dict):
            for k, v in obj.items():
                if k.lower() == key_name.lower():
                    return v
                ret = search_key(v, key_name)
                if ret is not None:
                    return ret
        elif isinstance(obj, list):
            for item in obj:
                ret = search_key(item, key_name)
                if ret is not None:
                    return ret
        return None

    val = search_key(data, "depth_unit_scale_m")
    if val is not None:
        depth_scale_m = float(val)

    if depth_scale_m is None:
        val = search_key(data, "depth_scale")
        if val is not None:
            val = float(val)
            if val < 0.1:
                depth_scale_m = val

    if depth_scale_m is None:
        val = search_key(data, "mm_per_unit")
        if val is not None:
            depth_scale_m = float(val) / 1000.0

    return intr, depth_scale_m


def get_intrinsics_and_scale(base_dir, root_dir, manual_intrinsics, manual_depth_scale_m):
    calib_path = find_calibration_json(base_dir, root_dir)

    if calib_path is not None:
        try:
            intr, depth_scale_m = parse_intrinsics_from_json(calib_path)

            if intr is None:
                intr = manual_intrinsics.copy()

            if depth_scale_m is None:
                depth_scale_m = manual_depth_scale_m

            print(f"[INFO] Calibration file: {calib_path}")
            print(f"[INFO] intrinsics={intr}, depth_scale_m={depth_scale_m}")
            return intr, depth_scale_m

        except Exception as e:
            print(f"[WARN] Failed to read calibration file: {calib_path} | {e}")

    intr = manual_intrinsics.copy()
    depth_scale_m = manual_depth_scale_m

    print(f"[WARN] Using manual intrinsics: {intr}")
    print(f"[WARN] Using manual depth_scale_m: {depth_scale_m}")

    return intr, depth_scale_m


def depth_to_vis(depth_masked):
    """
    Convert segmented depth image to an 8-bit visualization image.
    Only valid non-zero depth pixels are normalized.
    """
    vis = np.zeros_like(depth_masked, dtype=np.uint8)
    valid = depth_masked > 0

    if not np.any(valid):
        return vis

    vals = depth_masked[valid].astype(np.float32)
    vmin = float(vals.min())
    vmax = float(vals.max())

    if vmax <= vmin:
        vis[valid] = 255
        return vis

    scaled = (depth_masked.astype(np.float32) - vmin) / (vmax - vmin)
    scaled = np.clip(scaled, 0, 1)

    vis = (scaled * 255).astype(np.uint8)
    vis[~valid] = 0

    return vis


def make_segmented_rgb(rgb_img, mask_bin, background="black"):
    if background == "white":
        out = np.full_like(rgb_img, 255)
    else:
        out = np.zeros_like(rgb_img)

    out[mask_bin] = rgb_img[mask_bin]
    return out


def write_ply_xyzrgb_binary(ply_path, xyz, rgb):
    ply_path = Path(ply_path)
    ply_path.parent.mkdir(parents=True, exist_ok=True)

    n = xyz.shape[0]

    verts = np.empty(
        n,
        dtype=[
            ("x", "<f4"),
            ("y", "<f4"),
            ("z", "<f4"),
            ("red", "u1"),
            ("green", "u1"),
            ("blue", "u1"),
        ],
    )

    verts["x"] = xyz[:, 0].astype(np.float32)
    verts["y"] = xyz[:, 1].astype(np.float32)
    verts["z"] = xyz[:, 2].astype(np.float32)
    verts["red"] = rgb[:, 0].astype(np.uint8)
    verts["green"] = rgb[:, 1].astype(np.uint8)
    verts["blue"] = rgb[:, 2].astype(np.uint8)

    header = (
        "ply\n"
        "format binary_little_endian 1.0\n"
        f"element vertex {n}\n"
        "property float x\n"
        "property float y\n"
        "property float z\n"
        "property uchar red\n"
        "property uchar green\n"
        "property uchar blue\n"
        "end_header\n"
    )

    with open(ply_path, "wb") as f:
        f.write(header.encode("ascii"))
        f.write(verts.tobytes())


def make_segmented_pointcloud(
    depth_img,
    rgb_img,
    intr,
    depth_scale_m,
    min_depth_m,
    max_depth_m
):
    h, w = depth_img.shape[:2]

    if rgb_img.shape[:2] != (h, w):
        raise ValueError("RGB and depth image sizes are inconsistent.")

    fx, fy, cx, cy = intr["fx"], intr["fy"], intr["cx"], intr["cy"]

    depth_m = depth_img.astype(np.float32) * float(depth_scale_m)
    valid = (depth_m >= min_depth_m) & (depth_m <= max_depth_m)

    if not np.any(valid):
        return (
            np.zeros((0, 3), dtype=np.float32),
            np.zeros((0, 3), dtype=np.uint8),
        )

    v, u = np.where(valid)

    z = depth_m[v, u]
    x = (u.astype(np.float32) - float(cx)) * z / float(fx)
    y = (v.astype(np.float32) - float(cy)) * z / float(fy)

    xyz = np.stack([x, y, z], axis=1).astype(np.float32)

    # cv2 reads RGB image as BGR. Convert BGR to RGB for PLY color.
    rgb = rgb_img[v, u][:, ::-1].copy().astype(np.uint8)

    return xyz, rgb


def process_one_folder_pair(
    root_dir,
    lvl1,
    lvl2,
    manual_intrinsics,
    manual_depth_scale_m,
    min_depth_m,
    max_depth_m,
    auto_resize_rgb_to_depth,
    skip_existing,
    rgb_background
):
    base_dir = Path(root_dir) / lvl1 / lvl2

    rgb_dir = base_dir / "rgb_pngs"
    depth_dir = base_dir / "depth_pngs"
    yolo_out = base_dir / "yolo_out"
    mask_dir = yolo_out / "mask"

    if not rgb_dir.exists():
        print(f"[WARN] Missing directory: {rgb_dir}")
        return 0, 0, 0, 0

    if not depth_dir.exists():
        print(f"[WARN] Missing directory: {depth_dir}")
        return 0, 0, 0, 0

    if not mask_dir.exists():
        print(f"[WARN] Missing directory: {mask_dir}")
        return 0, 0, 0, 0

    out_rgb_dir = yolo_out / "segmented_rgb"
    out_depth_dir = yolo_out / "segmented_depth"
    out_ply_dir = yolo_out / "segmented_ply"

    out_rgb_dir.mkdir(parents=True, exist_ok=True)
    out_depth_dir.mkdir(parents=True, exist_ok=True)
    out_ply_dir.mkdir(parents=True, exist_ok=True)

    rgb_map = get_image_map(rgb_dir)
    depth_map = get_image_map(depth_dir)
    mask_map = get_image_map(mask_dir)

    common_depth_mask = sorted(set(depth_map.keys()) & set(mask_map.keys()))
    common_all = sorted(set(rgb_map.keys()) & set(depth_map.keys()) & set(mask_map.keys()))

    if len(common_depth_mask) == 0:
        print(f"[WARN] No matched depth and mask files in: {base_dir}")
        return 0, 0, 0, 0

    intr, depth_scale_m = get_intrinsics_and_scale(
        base_dir=base_dir,
        root_dir=root_dir,
        manual_intrinsics=manual_intrinsics,
        manual_depth_scale_m=manual_depth_scale_m
    )

    print(f"[INFO] Processing folder: {base_dir}")
    print(f"[INFO] Matched depth + mask files: {len(common_depth_mask)}")
    print(f"[INFO] Matched rgb + depth + mask files: {len(common_all)}")

    cnt_rgb = 0
    cnt_depth = 0
    cnt_ply = 0
    cnt_empty_ply = 0

    # 1) Save segmented depth visualization images.
    for stem in common_depth_mask:
        depth_path = depth_map[stem]
        mask_path = mask_map[stem]
        out_depth = out_depth_dir / f"{stem}.png"

        if skip_existing and out_depth.exists():
            cnt_depth += 1
            continue

        depth = cv_imread(depth_path, cv2.IMREAD_UNCHANGED)
        mask = cv_imread(mask_path, cv2.IMREAD_GRAYSCALE)

        if depth is None:
            print(f"[WARN] Failed to read depth image: {depth_path}")
            continue

        if mask is None:
            print(f"[WARN] Failed to read mask image: {mask_path}")
            continue

        if depth.ndim == 3:
            depth = depth[:, :, 0]

        if mask.shape[:2] != depth.shape[:2]:
            mask = cv2.resize(
                mask,
                (depth.shape[1], depth.shape[0]),
                interpolation=cv2.INTER_NEAREST
            )

        mask_bin = mask > 127
        depth_masked = np.where(mask_bin, depth, 0).astype(depth.dtype)

        vis = depth_to_vis(depth_masked)
        cv_imwrite(out_depth, vis)
        cnt_depth += 1

    # 2) Save segmented RGB images and segmented point clouds.
    for stem in common_all:
        rgb_path = rgb_map[stem]
        depth_path = depth_map[stem]
        mask_path = mask_map[stem]

        out_rgb = out_rgb_dir / f"{stem}.png"
        out_ply = out_ply_dir / f"{stem}.ply"

        if skip_existing and out_rgb.exists() and out_ply.exists():
            cnt_rgb += 1
            cnt_ply += 1
            continue

        rgb = cv_imread(rgb_path, cv2.IMREAD_COLOR)
        depth = cv_imread(depth_path, cv2.IMREAD_UNCHANGED)
        mask = cv_imread(mask_path, cv2.IMREAD_GRAYSCALE)

        if rgb is None:
            print(f"[WARN] Failed to read RGB image: {rgb_path}")
            continue

        if depth is None:
            print(f"[WARN] Failed to read depth image: {depth_path}")
            continue

        if mask is None:
            print(f"[WARN] Failed to read mask image: {mask_path}")
            continue

        if depth.ndim == 3:
            depth = depth[:, :, 0]

        h, w = depth.shape[:2]

        if mask.shape[:2] != (h, w):
            mask = cv2.resize(mask, (w, h), interpolation=cv2.INTER_NEAREST)

        if rgb.shape[:2] != (h, w):
            if auto_resize_rgb_to_depth:
                rgb = cv2.resize(rgb, (w, h), interpolation=cv2.INTER_LINEAR)
            else:
                print(f"[WARN] RGB and depth image sizes are inconsistent. Skipped: {stem}")
                continue

        mask_bin = mask > 127

        if (not skip_existing) or (not out_rgb.exists()):
            segmented_rgb = make_segmented_rgb(
                rgb_img=rgb,
                mask_bin=mask_bin,
                background=rgb_background
            )
            cv_imwrite(out_rgb, segmented_rgb)
            cnt_rgb += 1
        else:
            cnt_rgb += 1

        depth_masked = np.where(mask_bin, depth, 0).astype(depth.dtype)

        if (not skip_existing) or (not out_ply.exists()):
            xyz, rgb_pts = make_segmented_pointcloud(
                depth_img=depth_masked,
                rgb_img=rgb,
                intr=intr,
                depth_scale_m=depth_scale_m,
                min_depth_m=min_depth_m,
                max_depth_m=max_depth_m
            )

            if xyz.shape[0] == 0:
                cnt_empty_ply += 1
                print(f"[WARN] Empty point cloud: {out_ply}")
                continue

            write_ply_xyzrgb_binary(out_ply, xyz, rgb_pts)
            cnt_ply += 1
        else:
            cnt_ply += 1

    print(f"[INFO] Finished folder: {base_dir}")
    print(f"[INFO] segmented_rgb output count: {cnt_rgb}")
    print(f"[INFO] segmented_depth output count: {cnt_depth}")
    print(f"[INFO] segmented_ply output count: {cnt_ply}")
    print(f"[INFO] empty point cloud count: {cnt_empty_ply}")
    print("-" * 80)

    return cnt_rgb, cnt_depth, cnt_ply, cnt_empty_ply


def main():
    args = parse_args()

    root_dir = Path(args.root_dir).resolve()
    excel_path = Path(args.excel_path).resolve() if args.excel_path else root_dir / "data.xlsx"

    if not root_dir.exists():
        raise FileNotFoundError(f"root_dir does not exist: {root_dir}")

    if not excel_path.exists():
        raise FileNotFoundError(f"excel_path does not exist: {excel_path}")

    manual_intrinsics = {
        "fx": args.fx,
        "fy": args.fy,
        "cx": args.cx,
        "cy": args.cy,
    }

    pairs = parse_excel_pairs(excel_path)

    print("=" * 100)
    print(f"[INFO] root_dir = {root_dir}")
    print(f"[INFO] excel_path = {excel_path}")
    print(f"[INFO] folder pairs parsed from Excel: {len(pairs)}")
    print(f"[INFO] manual_intrinsics = {manual_intrinsics}")
    print(f"[INFO] manual_depth_scale_m = {args.depth_scale_m}")
    print(f"[INFO] valid depth range = [{args.min_depth_m}, {args.max_depth_m}] m")
    print(f"[INFO] auto_resize_rgb_to_depth = {args.auto_resize_rgb_to_depth}")
    print(f"[INFO] skip_existing = {args.skip_existing}")
    print(f"[INFO] rgb_background = {args.rgb_background}")
    print("=" * 100)

    total_rgb = 0
    total_depth = 0
    total_ply = 0
    total_empty_ply = 0
    total_folder = 0

    for lvl1, lvl2 in pairs:
        cnt_rgb, cnt_depth, cnt_ply, cnt_empty_ply = process_one_folder_pair(
            root_dir=root_dir,
            lvl1=lvl1,
            lvl2=lvl2,
            manual_intrinsics=manual_intrinsics,
            manual_depth_scale_m=args.depth_scale_m,
            min_depth_m=args.min_depth_m,
            max_depth_m=args.max_depth_m,
            auto_resize_rgb_to_depth=args.auto_resize_rgb_to_depth,
            skip_existing=args.skip_existing,
            rgb_background=args.rgb_background
        )

        if cnt_rgb > 0 or cnt_depth > 0 or cnt_ply > 0 or cnt_empty_ply > 0:
            total_folder += 1

        total_rgb += cnt_rgb
        total_depth += cnt_depth
        total_ply += cnt_ply
        total_empty_ply += cnt_empty_ply

    print("=" * 100)
    print("[DONE]")
    print(f"Processed folders: {total_folder}")
    print(f"Total segmented_rgb files: {total_rgb}")
    print(f"Total segmented_depth files: {total_depth}")
    print(f"Total segmented_ply files: {total_ply}")
    print(f"Total empty point clouds: {total_empty_ply}")
    print("=" * 100)


if __name__ == "__main__":
    main()