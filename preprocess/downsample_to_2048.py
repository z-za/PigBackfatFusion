"""
按 Excel 结构批量处理 yolo_out/segmented_ply 下的点云：
- 单数列第1行：一级文件夹名
- 单数列其余行：二级文件夹名

处理流程:
    0. 去非有限点
    1. 统计离群点去除（仅去除明显离群点）
    2. 体素下采样到“接近”8000
    3. 严格整理到8000点
    4. FPS 下采样到 2048 点
"""

import os
import sys
import glob
import argparse
import traceback
import numpy as np
import open3d as o3d
from openpyxl import load_workbook


def build_pcd(points, colors=None, normals=None):
    pcd = o3d.geometry.PointCloud()
    pcd.points = o3d.utility.Vector3dVector(points)
    if colors is not None and len(colors) == len(points):
        pcd.colors = o3d.utility.Vector3dVector(colors)
    if normals is not None and len(normals) == len(points):
        pcd.normals = o3d.utility.Vector3dVector(normals)
    return pcd


def get_attrs(pcd):
    pts = np.asarray(pcd.points)
    cols = np.asarray(pcd.colors) if pcd.has_colors() else None
    nors = np.asarray(pcd.normals) if pcd.has_normals() else None
    return pts, cols, nors


def select_by_mask(pcd, mask):
    pts, cols, nors = get_attrs(pcd)
    pts2 = pts[mask]
    cols2 = cols[mask] if cols is not None else None
    nors2 = nors[mask] if nors is not None else None
    return build_pcd(pts2, cols2, nors2)


def select_by_indices_np(pcd, indices):
    pts, cols, nors = get_attrs(pcd)
    pts2 = pts[indices]
    cols2 = cols[indices] if cols is not None else None
    nors2 = nors[indices] if nors is not None else None
    return build_pcd(pts2, cols2, nors2)


def remove_non_finite_points_np(pcd):
    pts, cols, nors = get_attrs(pcd)
    if len(pts) == 0:
        return pcd, 0

    mask = np.isfinite(pts).all(axis=1)
    removed = int((~mask).sum())
    return select_by_mask(pcd, mask), removed


def estimate_spacing(pcd):
    if len(pcd.points) < 2:
        return None
    dists = np.asarray(pcd.compute_nearest_neighbor_distance(), dtype=np.float64)
    dists = dists[np.isfinite(dists)]
    dists = dists[dists > 0]
    if len(dists) == 0:
        return None
    return float(np.median(dists))


def statistical_outlier_removal(pcd, nb_neighbors=30, std_ratio=2.0):
    """
    只保留明显离群点去除
    std_ratio 越小越严格
    """
    if len(pcd.points) == 0:
        return pcd, 0
    filtered, _ = pcd.remove_statistical_outlier(
        nb_neighbors=nb_neighbors,
        std_ratio=std_ratio
    )
    removed = len(pcd.points) - len(filtered.points)
    return filtered, removed


def fps_sample_indices(points, num_samples):
    """
    纯 numpy 的 FPS
    points: (N, 3)
    return: (num_samples,) 索引
    """
    points = np.asarray(points, dtype=np.float32)
    n = points.shape[0]

    if n == 0:
        raise ValueError("空点云，无法执行FPS。")
    if num_samples <= 0:
        raise ValueError("num_samples 必须 > 0。")

    if n == num_samples:
        return np.arange(n, dtype=np.int64)

    if n < num_samples:
        # 点数不足时补点
        base = np.arange(n, dtype=np.int64)
        extra = np.random.choice(n, size=num_samples - n, replace=True)
        return np.concatenate([base, extra], axis=0)

    selected = np.empty((num_samples,), dtype=np.int64)

    # 初始点：取离质心最远点
    centroid = np.mean(points, axis=0, keepdims=True)
    dist2center = np.sum((points - centroid) ** 2, axis=1)
    farthest = int(np.argmax(dist2center))

    min_dist = np.full((n,), np.inf, dtype=np.float32)

    for i in range(num_samples):
        selected[i] = farthest
        cur = points[farthest:farthest + 1]
        dist = np.sum((points - cur) ** 2, axis=1)
        min_dist = np.minimum(min_dist, dist)
        farthest = int(np.argmax(min_dist))

    return selected


def fps_downsample_pcd(pcd, num_samples=2048):
    pts, cols, nors = get_attrs(pcd)
    if len(pts) == 0:
        return pcd, None
    idx = fps_sample_indices(pts, num_samples)
    out = select_by_indices_np(pcd, idx)
    return out, idx


def ensure_exact_points_pcd(pcd, target_points):
    """
    将点云严格整理到 target_points：
    - 大于 target_points: 用 FPS 压到 target_points
    - 小于 target_points: 随机重复补点
    """
    pts, cols, nors = get_attrs(pcd)
    n = len(pts)

    if n == 0:
        return pcd

    if n == target_points:
        return pcd

    if n > target_points:
        idx = fps_sample_indices(pts, target_points)
        return select_by_indices_np(pcd, idx)

    # n < target_points: 补点
    extra = np.random.choice(n, size=target_points - n, replace=True)
    idx = np.concatenate([np.arange(n, dtype=np.int64), extra], axis=0)
    return select_by_indices_np(pcd, idx)


def voxel_downsample_to_target(pcd, target_points=8000, min_points=2048, max_iter=20):
    """
    用二分搜索体素大小，使体素下采样后的点数尽量接近 target_points，
    且不少于 min_points。
    返回:
        pcd_ds, voxel_size, out_n
    """
    n0 = len(pcd.points)
    if n0 == 0:
        return pcd, 0.0, 0

    if n0 <= target_points:
        return pcd, 0.0, n0

    pts = np.asarray(pcd.points)
    bbox = pts.max(axis=0) - pts.min(axis=0)
    diag = float(np.linalg.norm(bbox))

    spacing = estimate_spacing(pcd)
    if spacing is None or spacing <= 0:
        spacing = max(diag / 2000.0, 1e-6)

    low = max(spacing * 0.5, 1e-6)
    high = max(spacing * 20.0, low * 2.0)

    best_pcd = None
    best_voxel = 0.0
    best_diff = float("inf")
    best_n = n0

    for _ in range(max_iter):
        mid = (low + high) / 2.0
        ds = pcd.voxel_down_sample(mid)
        m = len(ds.points)

        if m >= min_points:
            diff = abs(m - target_points)
            if diff < best_diff:
                best_diff = diff
                best_pcd = ds
                best_voxel = mid
                best_n = m

        if m > target_points:
            low = mid
        else:
            high = mid

    if best_pcd is None:
        return pcd, 0.0, n0

    return best_pcd, best_voxel, best_n


def print_bbox_info(tag, pcd):
    pts = np.asarray(pcd.points)
    if len(pts) == 0:
        print(f"[{tag}] empty")
        return
    xyz_min = pts.min(axis=0)
    xyz_max = pts.max(axis=0)
    size = xyz_max - xyz_min
    print(f"[{tag}] min = {xyz_min}")
    print(f"[{tag}] max = {xyz_max}")
    print(f"[{tag}] size= {size}")


def process_one_ply(
    in_ply,
    out_ply,
    sor_nb_neighbors=30,
    sor_std_ratio=2.0,
    num_points=2048,
    pre_fps_target=8000,
    pre_fps_max_iter=20
):
    print("=" * 100)
    print(f"[FILE] {in_ply}")

    pcd_raw = o3d.io.read_point_cloud(in_ply)
    if len(pcd_raw.points) == 0:
        raise RuntimeError("点云为空或读取失败。")

    print(f"[INFO] 原始点数: {len(pcd_raw.points)}")
    print_bbox_info("RAW", pcd_raw)

    # Step 0: 去非有限点
    pcd, removed_nonfinite = remove_non_finite_points_np(pcd_raw)
    print(f"[STEP 0] 去非有限点 removed = {removed_nonfinite}, remain = {len(pcd.points)}")
    if len(pcd.points) == 0:
        raise RuntimeError("去非有限点后为空。")

    # Step 1: 仅做统计离群点去除
    pcd, removed_sor = statistical_outlier_removal(
        pcd,
        nb_neighbors=sor_nb_neighbors,
        std_ratio=sor_std_ratio
    )
    print(f"[STEP 1] 统计离群点去除 removed = {removed_sor}, remain = {len(pcd.points)}")
    if len(pcd.points) == 0:
        raise RuntimeError("统计离群点去除后为空。")

    print_bbox_info("AFTER_SOR", pcd)

    # Step 2: 先体素下采样到接近 8000
    if len(pcd.points) > pre_fps_target:
        print(f"[STEP 2] 先做体素预下采样，目标点数约 {pre_fps_target} ...")
        pcd_prefps, voxel_size, prefps_n = voxel_downsample_to_target(
            pcd,
            target_points=pre_fps_target,
            min_points=num_points,
            max_iter=pre_fps_max_iter
        )
        print(f"[STEP 2] 体素预下采样完成: voxel_size={voxel_size:.6f}, remain={prefps_n}")
    else:
        pcd_prefps = pcd
        print(f"[STEP 2] 当前点数 {len(pcd.points)} <= {pre_fps_target}，跳过体素预下采样。")

    print_bbox_info("AFTER_VOXEL", pcd_prefps)

    # Step 3: 严格整理到 8000 点
    pcd_8000 = ensure_exact_points_pcd(pcd_prefps, pre_fps_target)
    print(f"[STEP 3] 严格整理到 {pre_fps_target} 点后，remain = {len(pcd_8000.points)}")

    # Step 4: 再做 FPS 到 2048
    print(f"[STEP 4] 开始 FPS 下采样到 {num_points} 点...")
    pcd_fps, _ = fps_downsample_pcd(pcd_8000, num_samples=num_points)
    print(f"[STEP 4] FPS 完成，remain = {len(pcd_fps.points)}")
    print_bbox_info("AFTER_FPS", pcd_fps)

    out_dir = os.path.dirname(out_ply)
    if out_dir != "":
        os.makedirs(out_dir, exist_ok=True)

    ok = o3d.io.write_point_cloud(
        out_ply,
        pcd_fps,
        write_ascii=False,
        compressed=False
    )
    if not ok:
        raise RuntimeError(f"保存失败: {out_ply}")

    print(f"[DONE] 已保存: {out_ply}")


def cell_to_str(v, pad_len=6):
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None

    if isinstance(v, float):
        if v.is_integer():
            s = str(int(v))
        else:
            s = s.rstrip("0").rstrip(".")

    if s.isdigit() and pad_len is not None:
        s = s.zfill(pad_len)

    return s


def parse_excel_pairs(excel_path):
    """
    Excel 规则：
    - 单数列第1行：一级文件夹
    - 单数列第2行开始：二级文件夹
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    pairs = []
    seen = set()

    for col in range(1, ws.max_column + 1, 2):
        lvl1 = cell_to_str(ws.cell(row=1, column=col).value, pad_len=6)
        if not lvl1:
            continue

        for row in range(2, ws.max_row + 1):
            lvl2 = cell_to_str(ws.cell(row=row, column=col).value, pad_len=6)
            if not lvl2:
                continue

            key = (lvl1, lvl2)
            if key not in seen:
                seen.add(key)
                pairs.append(key)

    return pairs


def main():
    parser = argparse.ArgumentParser(
        description="按 Excel 批量处理 yolo_out/segmented_ply 下的点云，并输出到同级 yolo_out/2048 文件夹"
    )
    parser.add_argument(
        "--root_dir",
        type=str,
        required=True,
        help="Root directory of the dataset, e.g., F:/data"
    )
    parser.add_argument(
        "--excel_path",
        type=str,
        default=None,
        help="Path to data.xlsx. If not provided, root_dir/data.xlsx will be used."
    )

    parser.add_argument("--sor_nb_neighbors", type=int, default=30)
    parser.add_argument("--sor_std_ratio", type=float, default=3.0)

    parser.add_argument("--num_points", type=int, default=2048)

    parser.add_argument(
        "--pre_fps_target",
        type=int,
        default=8000,
        help="FPS前先整理到的目标点数，默认8000"
    )
    parser.add_argument(
        "--pre_fps_max_iter",
        type=int,
        default=20,
        help="体素下采样二分搜索迭代次数，默认20"
    )

    args = parser.parse_args()

    root_dir = os.path.abspath(args.root_dir)
    excel_path = os.path.abspath(args.excel_path) if args.excel_path else os.path.join(root_dir, "data.xlsx")

    if not os.path.isfile(excel_path):
        print(f"[ERROR] 未找到 Excel 文件: {excel_path}")
        sys.exit(1)

    print(f"[INFO] root_dir    = {root_dir}")
    print(f"[INFO] excel_path = {excel_path}")

    folder_pairs = parse_excel_pairs(excel_path)
    if len(folder_pairs) == 0:
        print("[ERROR] Excel 中没有读到有效的一级/二级文件夹对。")
        sys.exit(1)

    print(f"[INFO] Excel 共读到 {len(folder_pairs)} 个文件夹对")

    total_folders = 0
    total_files = 0
    success_files = 0
    failed_files = 0

    missing_folder_count = 0
    missing_segmented_ply_count = 0
    empty_segmented_ply_count = 0

    for lvl1, lvl2 in folder_pairs:
        sample_dir = os.path.join(root_dir, lvl1, lvl2)
        in_dir = os.path.join(sample_dir, "yolo_out", "segmented_ply")
        out_dir = os.path.join(sample_dir, "yolo_out", "2048")

        print("\n" + "#" * 120)
        print(f"[FOLDER] {lvl1} / {lvl2}")
        print(f"[PATH] sample_dir = {sample_dir}")

        if not os.path.isdir(sample_dir):
            print(f"[WARN] 文件夹不存在，跳过: {sample_dir}")
            missing_folder_count += 1
            continue

        if not os.path.isdir(in_dir):
            print(f"[WARN] 未找到 yolo_out/segmented_ply，跳过: {in_dir}")
            missing_segmented_ply_count += 1
            continue

        ply_files = sorted(glob.glob(os.path.join(in_dir, "*.ply")))
        if len(ply_files) == 0:
            print(f"[WARN] yolo_out/segmented_ply 下没有 .ply 文件，跳过: {in_dir}")
            empty_segmented_ply_count += 1
            continue

        os.makedirs(out_dir, exist_ok=True)
        total_folders += 1
        total_files += len(ply_files)

        print(f"[INFO] 找到 {len(ply_files)} 个 PLY 文件")
        print(f"[INFO] 输出目录: {out_dir}")

        for in_ply in ply_files:
            base_name = os.path.basename(in_ply)
            out_ply = os.path.join(out_dir, base_name)

            try:
                process_one_ply(
                    in_ply=in_ply,
                    out_ply=out_ply,
                    sor_nb_neighbors=args.sor_nb_neighbors,
                    sor_std_ratio=args.sor_std_ratio,
                    num_points=args.num_points,
                    pre_fps_target=args.pre_fps_target,
                    pre_fps_max_iter=args.pre_fps_max_iter
                )
                success_files += 1
            except Exception as e:
                failed_files += 1
                print(f"[ERROR] 处理失败: {in_ply}")
                print(f"[ERROR] {repr(e)}")
                traceback.print_exc()

    skipped_total = (
        missing_folder_count +
        missing_segmented_ply_count +
        empty_segmented_ply_count
    )

    print("\n" + "=" * 120)
    print("[SUMMARY]")
    print(f"有效文件夹数                      : {total_folders}")
    print(f"缺失样本文件夹数                  : {missing_folder_count}")
    print(f"缺失 segmented_ply 文件夹数       : {missing_segmented_ply_count}")
    print(f"segmented_ply 为空文件夹数        : {empty_segmented_ply_count}")
    print(f"跳过文件夹总数                    : {skipped_total}")
    print(f"总 PLY 文件数                     : {total_files}")
    print(f"成功处理文件数                    : {success_files}")
    print(f"失败处理文件数                    : {failed_files}")
    print("=" * 120)


if __name__ == "__main__":
    main()